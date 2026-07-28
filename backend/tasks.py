import threading
import uuid
from PIL import Image
import os, io, json, re, tempfile, shutil, zipfile
from datetime import datetime
from app import logger

import_tasks_store = {}
def get_drive_file_id(drive_link):
    patterns = [
        r'/folders/([a-zA-Z0-9_-]+)',
        r'/file/d/([a-zA-Z0-9_-]+)',
        r'id=([a-zA-Z0-9_-]+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, drive_link)
        if match:
            return match.group(1)
    return None

_drive_folder_cache = {}

# Guardrails on Google Drive folder pulls: admin-supplied links are a trusted-ish
# boundary, but an oversized/huge-file-count folder can still fill disk or hang
# the import worker. Cap what a single import will pull down.
MAX_DRIVE_FILES = 1000
MAX_DRIVE_TOTAL_BYTES = 1024 * 1024 * 1024  # 1 GB

# Guardrails on the local image ZIP upload — same spirit as the Drive caps
# above, plus zip-specific checks (encryption, path traversal, pathological
# archive structure) since this file comes straight from an admin's machine.
SUPPORTED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}
MAX_IMPORT_FILES = 2000
MAX_IMPORT_TOTAL_BYTES = 4 * 1024 * 1024 * 1024  # 4 GB
MAX_IMPORT_PATH_DEPTH = 10
MAX_IMPORT_FILENAME_LENGTH = 255


def validate_and_index_zip(zip_path):
    """Validate an uploaded image ZIP and return (index, extraction_dir).

    index: {IMAGE_CODE_UPPER: absolute_path_on_disk}
    extraction_dir: temp dir the ZIP was extracted into — caller must remove it.

    Raises ValueError with a clear message on any validation failure. Every
    entry in the archive is checked before a single byte is extracted.
    """
    if not zipfile.is_zipfile(zip_path):
        raise ValueError("Uploaded file is not a valid ZIP archive.")

    with zipfile.ZipFile(zip_path) as zf:
        infolist = zf.infolist()

        for zinfo in infolist:
            if zinfo.flag_bits & 0x1:
                raise ValueError(
                    f"ZIP contains an encrypted entry ('{zinfo.filename}') - "
                    "password-protected archives are not supported."
                )

        file_entries = [z for z in infolist if not z.is_dir()]

        if len(file_entries) > MAX_IMPORT_FILES:
            raise ValueError(
                f"ZIP contains {len(file_entries)} files, exceeding the "
                f"{MAX_IMPORT_FILES}-file import limit."
            )

        total_bytes = sum(z.file_size for z in file_entries)
        if total_bytes > MAX_IMPORT_TOTAL_BYTES:
            raise ValueError(
                f"ZIP's uncompressed contents total {total_bytes} bytes, "
                f"exceeding the {MAX_IMPORT_TOTAL_BYTES}-byte import limit."
            )

        extraction_dir = tempfile.mkdtemp(prefix='kj_import_')
        extraction_dir_real = os.path.realpath(extraction_dir)

        for zinfo in file_entries:
            normalized = os.path.normpath(zinfo.filename)
            if normalized.startswith('..') or os.path.isabs(normalized):
                shutil.rmtree(extraction_dir, ignore_errors=True)
                raise ValueError(
                    f"ZIP entry '{zinfo.filename}' resolves outside the archive "
                    "root - rejected (path traversal guard)."
                )

            depth = normalized.count(os.sep) + 1
            if depth > MAX_IMPORT_PATH_DEPTH:
                shutil.rmtree(extraction_dir, ignore_errors=True)
                raise ValueError(
                    f"ZIP entry '{zinfo.filename}' exceeds the maximum path "
                    f"depth of {MAX_IMPORT_PATH_DEPTH}."
                )

            if len(os.path.basename(normalized)) > MAX_IMPORT_FILENAME_LENGTH:
                shutil.rmtree(extraction_dir, ignore_errors=True)
                raise ValueError(
                    f"ZIP entry '{zinfo.filename}' has a filename longer than "
                    f"{MAX_IMPORT_FILENAME_LENGTH} characters."
                )

            target_path = os.path.realpath(os.path.join(extraction_dir, normalized))
            if not (target_path == extraction_dir_real
                    or target_path.startswith(extraction_dir_real + os.sep)):
                shutil.rmtree(extraction_dir, ignore_errors=True)
                raise ValueError(
                    f"ZIP entry '{zinfo.filename}' resolves outside the "
                    "extraction directory - rejected (path traversal guard)."
                )

        # Every entry passed every check - safe to extract now.
        zf.extractall(extraction_dir)

    index = {}
    duplicates = set()
    skipped_non_image = 0
    for root, dirs, files in os.walk(extraction_dir):
        for fname in files:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in SUPPORTED_IMAGE_EXTENSIONS:
                skipped_non_image += 1
                continue
            code = os.path.splitext(fname)[0].upper()
            path = os.path.join(root, fname)
            if code in index:
                duplicates.add(code)
            else:
                index[code] = path

    if duplicates:
        shutil.rmtree(extraction_dir, ignore_errors=True)
        raise ValueError(
            "ZIP contains duplicate image codes (same code, different files): "
            + ", ".join(sorted(duplicates))
        )

    logger.info(
        f"ZIP import indexed: {len(index)} images, {skipped_non_image} "
        f"non-image files skipped, from {len(file_entries)} archive entries."
    )
    return index, extraction_dir

def process_and_store_image(file_bytes, image_code, save_folder, upload_folder):
    img = Image.open(io.BytesIO(file_bytes))
    if img.mode in ('RGBA', 'LA', 'P'):
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'P':
            img = img.convert('RGBA')
        background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
        img = background

    img_ratio = img.size[0] / img.size[1]
    if img_ratio > 1.0:
        new_height = 1080
        new_width = int(new_height * img_ratio)
    else:
        new_width = 1080
        new_height = int(new_width / img_ratio)
    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
    left = (new_width - 1080) // 2
    top = (new_height - 1080) // 2
    img = img.crop((left, top, left + 1080, top + 1080))

    out = io.BytesIO()
    img.save(out, format='WEBP', quality=85)
    out.seek(0)

    folder_path = os.path.join(upload_folder, save_folder)
    os.makedirs(folder_path, exist_ok=True)
    save_filename = f"{int(datetime.now().timestamp())}_{image_code}.webp"
    save_path = os.path.join(folder_path, save_filename)
    with open(save_path, 'wb') as f:
        f.write(out.read())

    return f"/uploads/{save_folder}/{save_filename}"


def download_image_from_drive(folder_id, image_code, save_folder, upload_folder):
    try:
        import gdown
        if folder_id not in _drive_folder_cache:
            output_dir = os.path.join(tempfile.gettempdir(), f'kj_drive_{folder_id}')
            if os.path.exists(output_dir):
                shutil.rmtree(output_dir)
            os.makedirs(output_dir, exist_ok=True)
            gdown.download_folder(
                url=f"https://drive.google.com/drive/folders/{folder_id}",
                output=output_dir, quiet=False, use_cookies=False
            )

            file_count = 0
            total_bytes = 0
            for root, dirs, files in os.walk(output_dir):
                for fname in files:
                    file_count += 1
                    total_bytes += os.path.getsize(os.path.join(root, fname))
            if file_count > MAX_DRIVE_FILES or total_bytes > MAX_DRIVE_TOTAL_BYTES:
                shutil.rmtree(output_dir, ignore_errors=True)
                logger.warning(
                    f"Drive folder {folder_id} exceeded import limits "
                    f"({file_count} files, {total_bytes} bytes) - skipped"
                )
                _drive_folder_cache[folder_id] = {}
            else:
                file_index = {}
                for root, dirs, files in os.walk(output_dir):
                    for fname in files:
                        base = fname.rsplit('.', 1)[0].upper()
                        file_index[base] = os.path.join(root, fname)
                _drive_folder_cache[folder_id] = file_index

        file_index = _drive_folder_cache[folder_id]
        found_file = file_index.get(image_code.upper())
        if not found_file:
            return None

        with open(found_file, 'rb') as f:
            file_bytes = f.read()

        return process_and_store_image(file_bytes, image_code, save_folder, upload_folder)
    except Exception as e:
        logger.exception(f"Drive download error: {e}")
        return None

def resolve_image(image_code, save_folder, upload_folder, local_index, folder_id):
    """Resolve one image code to a saved, processed image path.

    Local ZIP index is checked first and wins if present; Drive is the
    fallback for any code not found locally.
    """
    if not image_code:
        return None
    local_path = local_index.get(image_code.upper()) if local_index else None
    if local_path:
        with open(local_path, 'rb') as f:
            return process_and_store_image(f.read(), image_code, save_folder, upload_folder)
    if folder_id:
        return download_image_from_drive(folder_id, image_code, save_folder, upload_folder)
    return None

def run_import_background(file_bytes, overwrite_images=False, zip_path=None):
    task_id = str(uuid.uuid4())
    import_tasks_store[task_id] = {'state': 'PENDING', 'progress': 0, 'status': 'Starting...', 'result': None}

    def background_worker():
        import_tasks_store[task_id]['state'] = 'PROGRESS'
        run_import_logic(task_id, file_bytes, overwrite_images, zip_path=zip_path)

    thread = threading.Thread(target=background_worker)
    thread.daemon = True
    thread.start()
    return task_id

def run_import_logic(task_id, file_bytes, overwrite_images, zip_path=None):
    import openpyxl
    from app import app, db
    from models import Segment, Category, Subcategory, Product

    extraction_dir = None
    try:
        with app.app_context():
            upload_folder = app.config['UPLOAD_FOLDER']
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            results = {'segments': 0, 'categories': 0, 'subcategories': 0, 'products': 0, 'errors': []}

            local_index = {}
            if zip_path:
                if task_id in import_tasks_store:
                    import_tasks_store[task_id]['status'] = 'Extracting image ZIP...'
                local_index, extraction_dir = validate_and_index_zip(zip_path)
                if task_id in import_tasks_store:
                    import_tasks_store[task_id]['status'] = f'Indexed {len(local_index)} images from ZIP'

            # Identify sheets flexibly
            sheet_map = {}
            for sn in wb.sheetnames:
                sn_norm = sn.upper().replace(' ', '')
                if 'SEGMENT' in sn_norm: sheet_map['segments'] = sn
                elif 'SUBCATEGORY' in sn_norm or 'SUBCAT' in sn_norm: sheet_map['subcategories'] = sn
                elif 'CATEGORY' in sn_norm: sheet_map['categories'] = sn
                elif 'PRODUCT' in sn_norm: sheet_map['products'] = sn

            # Count total rows for progress
            total_rows = 0
            for key in ['segments', 'categories', 'subcategories', 'products']:
                if key in sheet_map:
                    total_rows += wb[sheet_map[key]].max_row - 1
            
            done_rows = 0
            def update_progress(msg):
                nonlocal done_rows
                done_rows += 1
                pct = int((done_rows / max(total_rows, 1)) * 100)
                if task_id in import_tasks_store:
                    import_tasks_store[task_id]['progress'] = pct
                    import_tasks_store[task_id]['status'] = msg

            if not sheet_map:
                raise Exception("No valid inventory sheets found. Please ensure your Excel sheets are named correctly (e.g., SEGMENTS, CATEGORIES, SUBCATS, PRODUCTS).")

            # SHEET 1: SEGMENTS
            if task_id in import_tasks_store:
                import_tasks_store[task_id]['status'] = 'Importing Segments...'
            if 'segments' in sheet_map:
                ws = wb[sheet_map['segments']]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        update_progress("Skipping empty segment row")
                        continue
                    try:
                        seg_code = str(row[0]).strip()
                        seg_name = str(row[1]).strip() if row[1] else seg_code
                        img_code = str(row[2]).strip() if len(row) > 2 and row[2] else None
                        drive_link = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        disp_order = int(row[4]) if len(row) > 4 and row[4] else 0
                        is_active = str(row[5]).strip().upper() == 'YES' if len(row) > 5 and row[5] else True

                        existing = Segment.query.filter_by(segment_code=seg_code).first()
                        
                        img_path = None
                        should_download = True
                        if existing and existing.image_path and not overwrite_images:
                            should_download = False
                            
                        if should_download and img_code:
                            folder_id = None
                            if drive_link and 'PASTE' not in drive_link:
                                folder_id = get_drive_file_id(drive_link)
                            img_path = resolve_image(img_code, 'segments', upload_folder, local_index, folder_id)

                        if existing:
                            existing.name = seg_name
                            if img_path:
                                existing.image_path = img_path
                            existing.display_order = disp_order
                            existing.is_active = is_active
                        else:
                            db.session.add(Segment(name=seg_name, segment_code=seg_code, image_path=img_path, display_order=disp_order, is_active=is_active))
                        db.session.flush()
                        results['segments'] += 1
                        update_progress(f"Importing segment: {seg_name}")
                    except Exception as e:
                        results['errors'].append(f"Segment error: {str(e)}")
                        update_progress(f"Error in segment row")
                db.session.commit()

            # SHEET 2: CATEGORIES
            if task_id in import_tasks_store:
                import_tasks_store[task_id]['status'] = 'Importing Categories...'
            if 'categories' in sheet_map:
                ws = wb[sheet_map['categories']]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]: # Use cat_code as primary check
                        update_progress("Skipping category row")
                        continue
                    try:
                        seg_code = str(row[0]).strip()
                        cat_code = str(row[1]).strip()
                        cat_name = str(row[2]).strip() if row[2] else cat_code
                        img_code = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        gal_codes = [str(row[i]).strip() for i in range(4, 8) if len(row) > i and row[i]]
                        drive_link = str(row[8]).strip() if len(row) > 8 and row[8] else None
                        is_active = str(row[9]).strip().upper() == 'YES' if len(row) > 9 and row[9] else True

                        segment = Segment.query.filter_by(segment_code=seg_code).first()
                        if not segment:
                            results['errors'].append(f"Category '{cat_name}': segment '{seg_code}' not found")
                            update_progress(f"Segment not found for category")
                            continue

                        existing = Category.query.filter_by(category_code=cat_code).first()
                        
                        img_path = None
                        gallery_paths = []
                        should_download = True
                        if existing and existing.image_path and not overwrite_images:
                            should_download = False

                        if should_download:
                            folder_id = None
                            if drive_link and 'PASTE' not in drive_link:
                                folder_id = get_drive_file_id(drive_link)
                            if img_code:
                                img_path = resolve_image(img_code, 'categories', upload_folder, local_index, folder_id)
                            for gc in gal_codes:
                                gp = resolve_image(gc, 'categories', upload_folder, local_index, folder_id)
                                if gp:
                                    gallery_paths.append(gp)

                        if existing:
                            existing.name = cat_name
                            existing.segment_id = segment.id
                            if img_path:
                                existing.image_path = img_path
                            if gallery_paths:
                                existing.gallery_images = json.dumps(gallery_paths)
                            existing.is_active = is_active
                        else:
                            db.session.add(Category(name=cat_name, category_code=cat_code, segment_id=segment.id, image_path=img_path, gallery_images=json.dumps(gallery_paths) if gallery_paths else None, is_active=is_active))
                        db.session.flush()
                        results['categories'] += 1
                        update_progress(f"Importing category: {cat_name}")
                    except Exception as e:
                        results['errors'].append(f"Category error: {str(e)}")
                        update_progress(f"Error in category row")
                db.session.commit()

            # SHEET 3: SUBCATEGORIES
            if task_id in import_tasks_store:
                import_tasks_store[task_id]['status'] = 'Importing Subcategories...'
            if 'subcategories' in sheet_map:
                ws = wb[sheet_map['subcategories']]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]: # Use sub_code as primary check
                        update_progress("Skipping subcategory row")
                        continue
                    try:
                        cat_code = str(row[0]).strip()
                        sub_code = str(row[1]).strip()
                        sub_name = str(row[2]).strip() if row[2] else sub_code
                        img_code = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        gal_codes = [str(row[i]).strip() for i in range(4, 8) if len(row) > i and row[i]]
                        drive_link = str(row[8]).strip() if len(row) > 8 and row[8] else None
                        is_active = str(row[9]).strip().upper() == 'YES' if len(row) > 9 and row[9] else True

                        category = Category.query.filter_by(category_code=cat_code).first()
                        if not category:
                            results['errors'].append(f"Subcategory '{sub_name}': category '{cat_code}' not found")
                            update_progress(f"Category not found for subcategory")
                            continue

                        existing = Subcategory.query.filter_by(subcategory_code=sub_code).first()
                        
                        img_path = None
                        gallery_paths = []
                        should_download = True
                        if existing and existing.image_path and not overwrite_images:
                            should_download = False
                            
                        if should_download:
                            folder_id = None
                            if drive_link and 'PASTE' not in drive_link:
                                folder_id = get_drive_file_id(drive_link)
                            if img_code:
                                img_path = resolve_image(img_code, 'subcategories', upload_folder, local_index, folder_id)
                            for gc in gal_codes:
                                gp = resolve_image(gc, 'subcategories', upload_folder, local_index, folder_id)
                                if gp:
                                    gallery_paths.append(gp)

                        if existing:
                            existing.name = sub_name
                            existing.category_id = category.id
                            if img_path:
                                existing.image_path = img_path
                            if gallery_paths:
                                existing.gallery_images = json.dumps(gallery_paths)
                            existing.is_active = is_active
                        else:
                            db.session.add(Subcategory(name=sub_name, subcategory_code=sub_code, category_id=category.id, image_path=img_path, gallery_images=json.dumps(gallery_paths) if gallery_paths else None, is_active=is_active))
                        db.session.flush()
                        results['subcategories'] += 1
                        update_progress(f"Importing subcategory: {sub_name}")
                    except Exception as e:
                        results['errors'].append(f"Subcategory error: {str(e)}")
                        update_progress(f"Error in subcategory row")
                db.session.commit()

            # SHEET 4: PRODUCTS
            if task_id in import_tasks_store:
                import_tasks_store[task_id]['status'] = 'Importing Products...'
            if 'products' in sheet_map:
                ws = wb[sheet_map['products']]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]: # Use prod_code as primary check
                        update_progress("Skipping product row")
                        continue
                    try:
                        sub_code = str(row[0]).strip()
                        prod_code = str(row[1]).strip()
                        prod_name = str(row[2]).strip() if row[2] else prod_code
                        img_code = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        sec_codes = [str(row[i]).strip() for i in range(4, 8) if len(row) > i and row[i]]
                        drive_link = str(row[8]).strip() if len(row) > 8 and row[8] else None
                        details = str(row[9]).strip() if len(row) > 9 and row[9] else None
                        best_sell = str(row[10]).strip().upper() == 'YES' if len(row) > 10 and row[10] else False
                        is_assured = str(row[11]).strip().upper() == 'YES' if len(row) > 11 and row[11] else False
                        rating = float(row[12]) if len(row) > 12 and row[12] else 0.0
                        is_active = str(row[13]).strip().upper() == 'YES' if len(row) > 13 and row[13] else True

                        subcategory = Subcategory.query.filter_by(subcategory_code=sub_code).first()
                        if not subcategory:
                            results['errors'].append(f"Product '{prod_name}': subcategory '{sub_code}' not found")
                            update_progress("Subcategory not found for product")
                            continue

                        existing = Product.query.filter_by(product_code=prod_code).first()

                        primary_path = None
                        secondary_paths = []
                        should_download = True
                        if existing and existing.primary_image and not overwrite_images:
                            should_download = False

                        if should_download:
                            folder_id = None
                            if drive_link and 'PASTE' not in drive_link:
                                folder_id = get_drive_file_id(drive_link)
                            if img_code:
                                primary_path = resolve_image(img_code, 'products/primary', upload_folder, local_index, folder_id)
                            for sc in sec_codes:
                                sp = resolve_image(sc, 'products/gallery', upload_folder, local_index, folder_id)
                                if sp:
                                    secondary_paths.append(sp)

                        if existing:
                            existing.name = prod_name
                            existing.details = details
                            existing.subcategory_id = subcategory.id
                            existing.category_id = None
                            if primary_path:
                                existing.primary_image = primary_path
                            if secondary_paths:
                                existing.secondary_images = json.dumps(secondary_paths)
                            existing.is_best_selling = best_sell
                            existing.is_assured = is_assured
                            existing.rating = rating
                            existing.is_active = is_active
                        else:
                            db.session.add(Product(name=prod_name, product_code=prod_code, details=details, subcategory_id=subcategory.id, category_id=None, primary_image=primary_path, secondary_images=json.dumps(secondary_paths) if secondary_paths else None, is_best_selling=best_sell, is_assured=is_assured, rating=rating, is_active=is_active))
                        db.session.flush()
                        results['products'] += 1
                        update_progress(f"Importing product: {prod_name}")
                    except Exception as e:
                        results['errors'].append(f"Product error: {str(e)}")
                        update_progress("Error in product row")
                db.session.commit()

        logger.info(
            f"Import complete: {results['segments']} segments, {results['categories']} "
            f"categories, {results['subcategories']} subcategories, {results['products']} "
            f"products, {len(local_index)} images available from ZIP, "
            f"{len(results['errors'])} row errors."
        )
        if task_id in import_tasks_store:
            import_tasks_store[task_id]['state'] = 'SUCCESS'
            import_tasks_store[task_id]['progress'] = 100
            import_tasks_store[task_id]['result'] = results

    except Exception as e:
        if task_id in import_tasks_store:
            import_tasks_store[task_id]['state'] = 'FAILURE'
            import_tasks_store[task_id]['status'] = str(e)
        logger.exception("Import failed")
    finally:
        if extraction_dir and os.path.isdir(extraction_dir):
            shutil.rmtree(extraction_dir, ignore_errors=True)
        if zip_path and os.path.exists(zip_path):
            os.remove(zip_path)