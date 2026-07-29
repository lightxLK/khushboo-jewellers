"""
Convert the raw wide-format 'Final Product List Khushboo.xlsx' (35+ category
blocks laid out side-by-side across 4 sheets, one block per category: Product
Code / File Code / Sub Categories / Sub Categories Code / Name) into the
4-sheet template backend/tasks.py's /admin/import-excel expects:

    1 - Segments      (segment_code, segment_name, is_active)
    2 - Categories    (segment_code, category_code, category_name, is_active)
    3 - Subcategories (category_code, subcategory_code, subcategory_name, is_active)
    4 - Products      (sub_code, product_code, product_name, img_code, sec_1..4,
                        drive_link, details, best_sell, is_assured, rating, is_active)

img_code is set to the bare image filename stem (e.g. MPK08492) so it matches
by uppercased filename against the images ZIP (Product Images/All), per
tasks.py's validate_and_index_zip()/resolve_image() matching logic.

`details` is pulled from the Excel cell comment on each product's Name cell
(Type/Material/Design/Closure/Style spec), not the raw sheet value.

Run: python scripts/convert_catalog.py
"""
import os
import re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "Final Product List Khushboo.xlsx")
IMG_DIR = os.path.join(ROOT, "Product Images", "All")
OUT = os.path.join(ROOT, "backend", "Khushboo_Jewellers_Import.xlsx")

# Categories that are wearable jewellery -> segment JEW. Everything else -> POOJA.
JEWELLERY_CATEGORIES = {
    "Anklet (Categories)", "Waist Band", "Bracelet", "Gens Neck Chain",
    "Toe Ring", "Har", "Armlet (filigree)", "Wrist Watch",
}

SEGMENTS = [
    ("JEW", "Jewellery"),
    ("POOJA", "Pooja & Gift Items"),
]


def slugify_category_code(name, used):
    base = re.sub(r"\(.*", "", name)
    base = re.sub(r"[^A-Za-z0-9]", "", base).upper()
    code = "CAT-" + base[:6]
    n = 1
    final = code
    while final in used:
        n += 1
        final = f"{code}{n}"
    used.add(final)
    return final


def main():
    existing_images = set(os.listdir(IMG_DIR))
    wb_src = openpyxl.load_workbook(SRC, data_only=True)

    categories = []  # (segment_code, category_code, category_name)
    subcategories = []  # (category_code, subcategory_code, subcategory_name)
    products = []  # (sub_code, product_code, product_name, img_code, details)

    used_cat_codes = set()
    seen_subcodes = set()
    seq_by_category = {}
    missing_images = []
    missing_specs = []

    for ws in wb_src.worksheets:
        row1 = [c.value for c in ws[1]]
        row2 = [c.value for c in ws[2]]
        ncols = ws.max_column
        for c in range(ncols):
            if row2[c] != "Product Code":
                continue
            cat_name = None
            for k in range(c, -1, -1):
                if row1[k]:
                    cat_name = row1[k]
                    break
            if cat_name is None:
                continue

            clean_name = re.sub(r"\s*\(Categories\)\s*$", "", cat_name).strip()
            segment_code = "JEW" if cat_name in JEWELLERY_CATEGORIES else "POOJA"
            cat_code = slugify_category_code(cat_name, used_cat_codes)
            categories.append((segment_code, cat_code, clean_name))
            seq_by_category[cat_code] = 0

            for r in range(3, ws.max_row + 1):
                file_code = ws.cell(row=r, column=c + 2).value
                subcat_name = ws.cell(row=r, column=c + 3).value
                subcat_code = ws.cell(row=r, column=c + 4).value
                name_cell = ws.cell(row=r, column=c + 5)
                prod_name = name_cell.value
                if not file_code:
                    continue

                if subcat_code and subcat_code not in seen_subcodes:
                    seen_subcodes.add(subcat_code)
                    subcategories.append((cat_code, subcat_code, subcat_name or subcat_code))

                fname = f"MPK{int(file_code):05d}.JPG"
                if fname not in existing_images:
                    missing_images.append(fname)

                if not (name_cell.comment and name_cell.comment.text.strip()):
                    missing_specs.append((ws.title, prod_name, fname))
                spec = (name_cell.comment.text.strip() if name_cell.comment
                        else f"File Code: {int(file_code)}")

                seq_by_category[cat_code] += 1
                product_code = f"{cat_code}-{seq_by_category[cat_code]:03d}"
                products.append((
                    subcat_code, product_code, prod_name or product_code,
                    fname.rsplit(".", 1)[0],
                    spec,
                ))

    if missing_images:
        raise SystemExit(f"Missing {len(missing_images)} images, aborting: {missing_images[:10]}")
    if missing_specs:
        print(f"WARNING: {len(missing_specs)} products have no spec comment, "
              f"falling back to File Code note: {missing_specs[:10]}")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "1 - Segments"
    ws.append(["segment_code", "segment_name", "is_active"])
    for code, name in SEGMENTS:
        ws.append([code, name, "YES"])

    ws = wb_out.create_sheet("2 - Categories")
    ws.append(["segment_code", "category_code", "category_name", "is_active"])
    for seg, code, name in categories:
        ws.append([seg, code, name, "YES"])

    ws = wb_out.create_sheet("3 - Subcategories")
    ws.append(["category_code", "subcategory_code", "subcategory_name", "is_active"])
    for cat, code, name in subcategories:
        ws.append([cat, code, name, "YES"])

    ws = wb_out.create_sheet("4 - Products")
    ws.append(["sub_code", "product_code", "product_name", "img_code",
               "sec_1", "sec_2", "sec_3", "sec_4", "drive_link", "details",
               "best_sell", "is_assured", "rating", "is_active"])
    for sub_code, prod_code, name, img_code, details in products:
        ws.append([sub_code, prod_code, name, img_code,
                   None, None, None, None, None, details,
                   None, None, None, "YES"])

    wb_out.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Segments: {len(SEGMENTS)}  Categories: {len(categories)}  "
          f"Subcategories: {len(subcategories)}  Products: {len(products)}")


if __name__ == "__main__":
    main()
